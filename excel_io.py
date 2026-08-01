"""Excel read + write-back (spec section 4).

Columns are found by header name (row 1), not fixed position, so column
order in the sheet doesn't matter. Every write-back saves the workbook
immediately - that's what makes partial success survive a crash.
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import OrderGroup, ReturnOutcome, ReturnTask

log = logging.getLogger(__name__)

PENDING_STATUSES = {"to do", "pending"}

HEADERS = [
    "Platform", "Order ID", "Product/SKU", "Return Window",
    "Return ID", "Return Status", "Refund Amount", "Task Status",
    "Timestamp", "Log",
]


def _col_map(ws: Worksheet) -> dict[str, int]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    mapping = {}
    for cell in header_row:
        if cell.value:
            mapping[str(cell.value).strip()] = cell.column
    missing = [h for h in ("Platform", "Order ID", "Product/SKU", "Return Window",
                           "Return Status", "Task Status") if h not in mapping]
    if missing:
        raise ValueError(f"Excel sheet is missing required column(s): {missing}")
    return mapping


def _parse_window(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None  # unparseable -> agent will not guess, see ReturnTask.window_open


class ExcelStore:
    def __init__(self, path: str, sheet: str = "Tasks"):
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.path}")
        self.sheet_name = sheet

    def _open(self):
        try:
            wb = load_workbook(self.path)
        except PermissionError as e:
            raise PermissionError(
                f"Can't open {self.path} - close it in Excel first, then re-run."
            ) from e
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found in {self.path}")
        return wb, wb[self.sheet_name]

    def pending_by_order(self) -> list[OrderGroup]:
        wb, ws = self._open()
        cols = _col_map(ws)
        groups: dict[tuple[str, str], OrderGroup] = {}
        order = []  # preserve first-seen order
        for row in ws.iter_rows(min_row=2):
            row_idx = row[0].row
            status_cell = row[cols["Task Status"] - 1].value
            status = str(status_cell or "").strip().lower()
            if status not in PENDING_STATUSES:
                continue
            platform = str(row[cols["Platform"] - 1].value or "").strip().lower()
            order_id = str(row[cols["Order ID"] - 1].value or "").strip()
            sku = str(row[cols["Product/SKU"] - 1].value or "").strip()
            window = _parse_window(row[cols["Return Window"] - 1].value)
            if not platform or not order_id or not sku:
                log.warning("Row %d skipped: missing Platform/Order ID/SKU", row_idx)
                continue
            key = (platform, order_id)
            if key not in groups:
                groups[key] = OrderGroup(order_id=order_id, platform=platform)
                order.append(key)
            groups[key].items.append(ReturnTask(
                row_index=row_idx, platform=platform, order_id=order_id,
                sku=sku, return_window=window,
            ))
        wb.close()
        return [groups[k] for k in order]

    def mark_in_progress(self, task: ReturnTask) -> None:
        wb, ws = self._open()
        cols = _col_map(ws)
        ws.cell(row=task.row_index, column=cols["Task Status"]).value = "In Progress"
        wb.save(self.path)
        wb.close()

    def write_outcome(self, task: ReturnTask, outcome: ReturnOutcome) -> None:
        wb, ws = self._open()
        cols = _col_map(ws)
        ws.cell(row=task.row_index, column=cols["Return ID"]).value = outcome.return_id
        ws.cell(row=task.row_index, column=cols["Return Status"]).value = outcome.status.value
        ws.cell(row=task.row_index, column=cols["Refund Amount"]).value = outcome.refund_amount
        ws.cell(row=task.row_index, column=cols["Task Status"]).value = outcome.task_status
        if "Timestamp" in cols:
            ws.cell(row=task.row_index, column=cols["Timestamp"]).value = (
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
        if "Log" in cols:
            ws.cell(row=task.row_index, column=cols["Log"]).value = outcome.note
        wb.save(self.path)
        wb.close()
        log.info("Row %d (%s / %s) -> %s", task.row_index, task.order_id, task.sku,
                  outcome.status.value)


def create_template(path: str, sample_rows: list[list[str]]) -> str:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append(HEADERS)
    for row in sample_rows:
        ws.append(row)
    wb.save(p)
    return str(p)
